from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import os
import openpyxl
from openpyxl import Workbook
browser = webdriver.Chrome('C:\\Users\\DELL\\Downloads\\chromedriver')

def search(roll_no,a,b):
    browser.get('http://59.144.74.15/scheme17/studentresult/index.asp')
    elem = browser.find_element_by_name('RollNumber')
    elem.send_keys(roll_no)
    elem.send_keys(Keys.ENTER)
    data = browser.find_elements_by_xpath("/html/body/div[1]/table/tbody/tr[1]/td[2]/div")
    cgpi = browser.find_elements_by_xpath("/html/body/div[7]/table/tbody/tr[2]/td[3]")
    sgpi = browser.find_elements_by_xpath("/html/body/div[7]/table/tbody/tr[2]/td[1]")
    print(data[0].text, end=" ")
    print(cgpi[0].text)
    l=cgpi[0].text
    if '.' in l:
        k = l.index('.')
        no = l[k - 1:k + 4]
    else:
        no = (l[7:12])

    l2 = sgpi[0].text
    if '.' in l2:
        k2 = l2.index('.')
        no2 = l2[k2 - 1:k2 + 4]
    else:
        no2 = (l2[7:12])



    fname = 'Ranklist.xlsx'
    if(os.path.exists(fname)):
        workbook = openpyxl.load_workbook(fname)
        worksheet = workbook.get_sheet_by_name('Sheet')
    else:
        workbook = Workbook()
        worksheet = workbook.active

    worksheet.cell(row=a, column=b).value = data[0].text
    worksheet.cell(row=a, column=b + 1).value = (no)
    worksheet.cell(row=a, column=b + 2).value = (no2)
    workbook.save(fname)




def electrical():
    b = 17201
    c = 17294
    p = 2
    q = 1
    for i in range(b, c):
        search(i, p, q)
        p += 1


def cse_dd():
    a = "17mi"
    b = 501
    c = 561
    p = 2
    q = 1
    for i in range(b, c):
        s = a + str(i)
        p += 1
        if i==529 or i==532 or i==545 or i==560 or i==546 or i==555:
            continue
        else:
            search(s, p, q)




def cse_iiit():
    url='http://59.144.74.15/iiituna17/studentresult/index.asp'
    a = "iiitu17"
    b = 101
    c = 158
    p = 2
    q = 1
    for i in range(b, c):
        s = a + str(i)
        search(s, p, q)
        p += 1






a=int(input("IIIT Cse   - 1 "
            "Electrical - 2 "
            "CSE Dual   - 3 "))
if a == 1:
    cse_iiit()
elif a == 2:
    electrical()
elif a == 3:
    cse_dd()










